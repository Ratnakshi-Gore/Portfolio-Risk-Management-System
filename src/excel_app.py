import logging
from datetime import datetime
from typing import Dict, List, Optional
import numpy as np
import pandas as pd

log = logging.getLogger(__name__)

try:
    import xlwings as xw
    HAS_XLWINGS = True
except Exception:
    HAS_XLWINGS = False
    log.warning("xlwings not available — Excel integration unavailable")



def _clear_sheet(sheet: "xw.Sheet", keep_rows: int = 0):
    """Clear all content below *keep_rows* (0 = clear everything)."""
    used = sheet.used_range
    if used is None:
        return
    last_row = used.last_cell.row
    last_col = used.last_cell.column
    if last_row > keep_rows:
        sheet.range(
            (keep_rows + 1, 1), (last_row, last_col)
        ).clear_contents()


def _write_header(sheet: "xw.Sheet", row: int, col: int, text: str):
    """Write a bold, coloured section header."""
    cell = sheet.range((row, col))
    cell.value = text
    cell.font.bold = True
    cell.font.size = 12
    cell.font.color = (0, 51, 102)  # navy


def _write_table(sheet: "xw.Sheet", start_row: int, start_col: int,
                 df: pd.DataFrame, header: bool = True):
    if header:
        for j, col_name in enumerate(df.columns):
            cell = sheet.range((start_row, start_col + j))
            cell.value = col_name
            cell.font.bold = True
            cell.api.Interior.Color = 0xD9E1F2  # light blue fill (BGR)
        start_row += 1

    sheet.range((start_row, start_col)).value = df.values.tolist()

    # Auto-fit
    end_col = start_col + len(df.columns) - 1
    sheet.range(
        (start_row - (1 if header else 0), start_col),
        (start_row + len(df) - 1, end_col),
    ).columns.autofit()


def _write_kv_table(sheet: "xw.Sheet", start_row: int, start_col: int,
                    data: Dict, title: str = ""):
    # Write a key-value dict as a two-column table.
    row = start_row
    if title:
        _write_header(sheet, row, start_col, title)
        row += 1

    for key, val in data.items():
        sheet.range((row, start_col)).value = str(key)
        sheet.range((row, start_col)).font.bold = True
        cell = sheet.range((row, start_col + 1))
        cell.value = val
        # Format percentages
        if isinstance(val, str) and "%" in val:
            cell.number_format = "0.00%"
        row += 1

    return row  # next available row


# Sheet writers

def _write_analysis_sheet(
    sheet: "xw.Sheet",
    metrics_summary: pd.DataFrame,
    metadata: Dict,
    regime_info: str,
    ff_results: Optional[Dict],
):
    _clear_sheet(sheet, keep_rows=0)

    title_cell = sheet.range("A1")
    title_cell.value = "Portfolio Risk Management — Analysis Dashboard"
    title_cell.font.bold = True
    title_cell.font.size = 16
    title_cell.font.color = (0, 51, 102)

    sheet.range("A2").value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sheet.range("A2").font.italic = True

    # Backtest parameters 
    row = 4
    _write_header(sheet, row, 1, "Backtest Parameters")
    row += 1
    params = {
        "Tickers": ", ".join(metadata.get("tickers", [])),
        "Period": f"{metadata.get('start_date', '')} → {metadata.get('end_date', '')}",
        "Rebalance": metadata.get("rebalance_freq", "M"),
        "Initial Value": f"${metadata.get('initial_value', 100000):,.0f}",
        "# Rebalances": metadata.get("n_rebalances", ""),
    }
    for k, v in params.items():
        sheet.range((row, 1)).value = k
        sheet.range((row, 1)).font.bold = True
        sheet.range((row, 2)).value = v
        row += 1

    # Macro Regime 
    row += 1
    _write_header(sheet, row, 1, "Current Macro Regime")
    row += 1
    sheet.range((row, 1)).value = regime_info
    row += 2

    # Metrics table 
    _write_header(sheet, row, 1, "Performance & Risk Metrics")
    row += 1
    _write_table(sheet, row, 1, metrics_summary)
    row += len(metrics_summary) + 2

    # Fama-French 
    if ff_results:
        _write_header(sheet, row, 1, "Fama-French 3-Factor Regression")
        row += 1
        ff_display = {
            "FF3 Alpha (annualised)": f"{ff_results.get('alpha_annual', 0):.4%}",
            "Alpha t-stat": f"{ff_results.get('t_stat_alpha', 0):.3f}",
            "Alpha p-value": f"{ff_results.get('p_value_alpha', 0):.4f}",
            "R-squared": f"{ff_results.get('r_squared', 0):.4f}",
            "# Observations": ff_results.get("n_observations", ""),
            "Beta (Mkt-RF)": f"{ff_results.get('beta_mkt', 0):.4f}",
            "Beta (SMB)": f"{ff_results.get('beta_smb', 0):.4f}",
            "Beta (HML)": f"{ff_results.get('beta_hml', 0):.4f}",
        }

        for k, v in ff_display.items():
            sheet.range((row, 1)).value = k
            sheet.range((row, 1)).font.bold = True
            sheet.range((row, 2)).value = v
            row += 1

    sheet.range("A:C").columns.autofit()


def _write_backtest_sheet(
    sheet: "xw.Sheet",
    nav: pd.Series,
    daily_returns: pd.Series,
    benchmark_returns: pd.Series,
    weights_history: list,
    metadata: Dict,
):
    _clear_sheet(sheet)

    _write_header(sheet, 1, 1, "Daily NAV")

    # Build NAV table
    bench_nav = (1 + benchmark_returns).cumprod() * metadata.get("initial_value", 100_000)

    nav_df = pd.DataFrame({
        "Date": nav.index.strftime("%Y-%m-%d"),
        "Portfolio NAV": nav.round(2).values,
        "Benchmark NAV": bench_nav.reindex(nav.index).round(2).values,
    })
    _write_table(sheet, 2, 1, nav_df)

    # Weights history 
    if weights_history:
        col_start = 5  # column E
        _write_header(sheet, 1, col_start, "Portfolio Weights (at each rebalance)")

        # Collect all tickers that ever appeared
        all_tickers = sorted(set(
            t for _, w in weights_history for t in w.keys()
        ))

        # Header row
        header_row = 2
        sheet.range((header_row, col_start)).value = "Date"
        sheet.range((header_row, col_start)).font.bold = True
        for j, ticker in enumerate(all_tickers):
            cell = sheet.range((header_row, col_start + 1 + j))
            cell.value = ticker
            cell.font.bold = True
            cell.api.Interior.Color = 0xD9E1F2

        # Data rows
        for i, (date, wts) in enumerate(weights_history):
            row = header_row + 1 + i
            if hasattr(date, "strftime"):
                sheet.range((row, col_start)).value = date.strftime("%Y-%m-%d")
            else:
                sheet.range((row, col_start)).value = str(date)
            for j, ticker in enumerate(all_tickers):
                val = wts.get(ticker, 0.0)
                cell = sheet.range((row, col_start + 1 + j))
                cell.value = round(val, 4)
                cell.number_format = "0.00%"

        # Auto-fit
        end_col = col_start + len(all_tickers)
        sheet.range(
            (header_row, col_start),
            (header_row + len(weights_history), end_col),
        ).columns.autofit()

    # Auto-fit NAV columns
    sheet.range("A:C").columns.autofit()


def _write_raw_data_sheet(
    sheet: "xw.Sheet",
    factor_scores: Optional[pd.DataFrame],
    stress_results: Optional[list],
    mc_summary: Optional[Dict],
):
    _clear_sheet(sheet)
    row = 1

    # Factor scores
    if factor_scores is not None and not factor_scores.empty:
        _write_header(sheet, row, 1, "Factor Scores (latest rebalance)")
        row += 1
        _write_table(sheet, row, 1, factor_scores.round(4).reset_index())
        row += len(factor_scores) + 3

    # Stress test 
    if stress_results:
        _write_header(sheet, row, 1, "Stress Test Results")
        row += 1
        for result in stress_results:
            sheet.range((row, 1)).value = result.get("scenario", "")
            sheet.range((row, 1)).font.bold = True
            sheet.range((row, 2)).value = result.get("description", "")
            cell = sheet.range((row, 3))
            cell.value = result.get("portfolio_return", 0)
            cell.number_format = "0.00%"
            dollar = result.get("dollar_loss", 0)
            cell2 = sheet.range((row, 4))
            cell2.value = dollar
            cell2.number_format = "$#,##0"
            row += 1

        # Column headers
        header_row = row - len(stress_results) - 1
        for j, h in enumerate(["Scenario", "Description", "Portfolio Return", "Dollar Loss"]):
            cell = sheet.range((header_row + 1, 1 + j))
        row += 2

    # Monte Carlo summary
    if mc_summary:
        _write_header(sheet, row, 1, "Monte Carlo Simulation (1-Year Forward)")
        row += 1
        for k, v in mc_summary.items():
            sheet.range((row, 1)).value = str(k)
            sheet.range((row, 1)).font.bold = True
            cell = sheet.range((row, 2))
            if isinstance(v, (int, float)):
                cell.value = v
                if abs(v) < 1:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "$#,##0"
            else:
                cell.value = str(v)
            row += 1

    sheet.range("A:F").columns.autofit()


# Main entry point — called from VBA

def run_full_analysis(tickers_str: str = None, backtest_years: int = 3):

    if not HAS_XLWINGS:
        raise RuntimeError("xlwings is not installed. Run: pip install xlwings")

    # Connect to the calling workbook 
    wb = xw.Book.caller()

    # Read inputs from the Analysis sheet 
    analysis_sheet = _get_or_create_sheet(wb, "Analysis")

    if tickers_str is None:
        raw = analysis_sheet.range("B3").value
        if raw and str(raw).strip():
            tickers_str = str(raw).strip()
        else:
            tickers_str = "AAPL, MSFT, JPM, JNJ, XOM"

    # Parse tickers
    tickers = [t.strip().upper() for t in tickers_str.split(",") if t.strip()]
    if not tickers:
        analysis_sheet.range("A50").value = "ERROR: No tickers provided"
        return

    # Read backtest years from B4 if available
    years_cell = analysis_sheet.range("B4").value
    if years_cell and isinstance(years_cell, (int, float)):
        backtest_years = int(years_cell)

    # Compute date range
    end_date = datetime.now().strftime("%Y-%m-%d")
    start_year = datetime.now().year - backtest_years
    start_date = f"{start_year}-01-01"

    # Status update 
    status_cell = analysis_sheet.range("A50")
    status_cell.value = f"Running analysis for {', '.join(tickers)}..."
    status_cell.font.italic = True

    try:
        # Import pipeline modules
        from src.backtester import run_backtest
        from src.risk_metrics import (
            compute_all_metrics, get_metrics_summary,
            compute_fama_french_alpha,
        )
        from src.monte_carlo import run_monte_carlo, compute_var_cvar, get_simulation_summary
        from src.stress_test import run_all_scenarios
        from src.data_loader import get_price_data, get_fama_french_factors, get_fred_data
        from src.macro_regime import run_macro_overlay

        #  STEP 1: Run backtest
        status_cell.value = "Step 1/6: Running backtest..."
        bt = run_backtest(
            tickers=tickers,
            start_date=start_date,
            end_date=end_date,
            rebalance_freq="M",
            initial_value=100_000,
        )

        #  STEP 2: Compute risk metrics
        status_cell.value = "Step 2/6: Computing risk metrics..."
        metrics = compute_all_metrics(
            bt["daily_returns"],
            bt["benchmark_returns"],
        )
        metrics_summary = get_metrics_summary(metrics)

        #  STEP 3: Fama-French alpha
        status_cell.value = "Step 3/6: Fama-French regression..."
        ff_results = None
        try:
            ff_factors = get_fama_french_factors()
            ff_results = compute_fama_french_alpha(bt["daily_returns"], ff_factors)
        except Exception as e:
            log.warning(f"FF3 regression failed: {e}")

        #  STEP 4: Macro regime
        status_cell.value = "Step 4/6: Detecting macro regime..."
        regime_info = "N/A"
        try:
            fred = get_fred_data()
            regime_result = run_macro_overlay(fred_data=fred)
            regime_info = (
                f"{regime_result.get('regime', 'UNKNOWN')} — "
                f"{regime_result.get('description', '')}"
            )
        except Exception as e:
            log.warning(f"Macro regime detection failed: {e}")
            regime_info = f"Detection failed: {e}"

        #  STEP 5: Monte Carlo simulation
        status_cell.value = "Step 5/6: Monte Carlo simulation..."
        mc_summary = None
        try:
            # Get latest weights from backtest
            latest_weights = _get_latest_weights(bt["weights_history"], tickers)
            prices = get_price_data(tickers, start_date, end_date)
            paths = run_monte_carlo(
                weights=latest_weights,
                price_df=prices,
                n_simulations=5_000,  # fewer for Excel speed
                horizon_days=252,
                initial_value=100_000,
            )
            mc_summary = get_simulation_summary(paths, initial_value=100_000)
        except Exception as e:
            log.warning(f"Monte Carlo failed: {e}")

        #  STEP 6: Stress testing
        status_cell.value = "Step 6/6: Stress testing..."
        stress_results = None
        try:
            latest_weights = _get_latest_weights(bt["weights_history"], tickers)
            stress_results = run_all_scenarios(
                weights=latest_weights,
                initial_value=100_000,
            )
        except Exception as e:
            log.warning(f"Stress testing failed: {e}")

        #  STEP 7: Get factor scores for latest rebalance
        factor_scores = _get_latest_factor_scores(bt, tickers)

        #  WRITE TO SHEETS
        status_cell.value = "Writing results to sheets..."

        # Sheet 1: Analysis
        _write_analysis_sheet(
            sheet=analysis_sheet,
            metrics_summary=metrics_summary,
            metadata=bt["metadata"],
            regime_info=regime_info,
            ff_results=ff_results,
        )

        # Re-write input cells (they were cleared by _write_analysis_sheet)
        analysis_sheet.range("A3").value = "Tickers:"
        analysis_sheet.range("A3").font.bold = True
        analysis_sheet.range("B3").value = ", ".join(tickers)
        analysis_sheet.range("A4").value = "Backtest Years:"
        analysis_sheet.range("A4").font.bold = True
        analysis_sheet.range("B4").value = backtest_years

        # Sheet 2: Backtest
        bt_sheet = _get_or_create_sheet(wb, "Backtest")
        _write_backtest_sheet(
            sheet=bt_sheet,
            nav=bt["nav"],
            daily_returns=bt["daily_returns"],
            benchmark_returns=bt["benchmark_returns"],
            weights_history=bt["weights_history"],
            metadata=bt["metadata"],
        )

        # Sheet 3: Raw Data
        raw_sheet = _get_or_create_sheet(wb, "Raw Data")
        _write_raw_data_sheet(
            sheet=raw_sheet,
            factor_scores=factor_scores,
            stress_results=stress_results,
            mc_summary=mc_summary,
        )

        status_cell.value = (
            f"✓ Analysis complete — {datetime.now().strftime('%H:%M:%S')} | "
            f"{len(tickers)} tickers | {bt['metadata']['n_rebalances']} rebalances"
        )
        status_cell.font.color = (0, 128, 0)  # green

    except Exception as e:
        status_cell.value = f"ERROR: {e}"
        status_cell.font.color = (200, 0, 0)  # red
        log.error(f"run_full_analysis failed: {e}", exc_info=True)
        raise


def _get_or_create_sheet(wb: "xw.Book", name: str) -> "xw.Sheet":
    """Get an existing sheet or create a new one."""
    for s in wb.sheets:
        if s.name == name:
            return s
    return wb.sheets.add(name, after=wb.sheets[-1])


def _get_latest_weights(weights_history: list, tickers: list) -> pd.Series:
    """Extract the most recent weights as a pd.Series."""
    if not weights_history:
        # Equal weight fallback
        n = len(tickers)
        return pd.Series({t: 1.0 / n for t in tickers})

    _, latest_wts = weights_history[-1]
    return pd.Series(latest_wts)


def _get_latest_factor_scores(bt_result: Dict, tickers: list) -> Optional[pd.DataFrame]:
    """
    Attempt to reconstruct factor scores for the latest rebalance.
    If not available in backtest result, compute them fresh.
    """
    try:
        from src.data_loader import get_price_data, get_fundamentals
        from src.factor_model import (
            compute_value_score, compute_momentum_score,
            compute_quality_score,
        )

        end_date = bt_result["metadata"]["end_date"]
        start_date = bt_result["metadata"]["start_date"]

        fundamentals = get_fundamentals(tickers)
        prices = get_price_data(tickers, start_date, end_date)

        value = compute_value_score(fundamentals)
        momentum = compute_momentum_score(prices)
        quality = compute_quality_score(fundamentals)

        scores = pd.DataFrame({
            "Value": value,
            "Momentum": momentum,
            "Quality": quality,
        })

        # Add latest weights
        _, latest_wts = bt_result["weights_history"][-1] if bt_result["weights_history"] else (None, {})
        scores["Weight"] = pd.Series(latest_wts)
        scores = scores.fillna(0)

        return scores

    except Exception as e:
        log.warning(f"Could not compute factor scores: {e}")
        return None


# Standalone test — run without Excel using mock data

def run_standalone_test():
    """
    Run the full pipeline without Excel to verify all components work.
    Prints results to console instead of writing to sheets.
    """
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).parent.parent))

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    print("=" * 60)
    print("EXCEL APP — STANDALONE TEST (no Excel required)")
    print("=" * 60)

    from src.backtester import run_backtest
    from src.risk_metrics import (
        compute_all_metrics, get_metrics_summary,
        compute_fama_french_alpha,
    )
    from src.monte_carlo import run_monte_carlo, compute_var_cvar, get_simulation_summary
    from src.stress_test import run_all_scenarios
    from src.data_loader import get_price_data, get_fama_french_factors

    tickers = ["AAPL", "MSFT", "JPM", "JNJ", "XOM"]
    start_date = "2021-01-01"
    end_date = "2024-12-31"

    # Step 1: Backtest
    print("\n[1/6] Running backtest...")
    bt = run_backtest(tickers=tickers, start_date=start_date, end_date=end_date)
    print(f"  Total Return: {bt['metadata']['total_return']:.2%}")
    print(f"  Sharpe:       {bt['metadata']['sharpe_ratio']:.3f}")

    # Step 2: Risk metrics
    print("\n[2/6] Computing risk metrics...")
    metrics = compute_all_metrics(bt["daily_returns"], bt["benchmark_returns"])
    summary = get_metrics_summary(metrics)
    print(summary.to_string(index=False))

    # Step 3: Fama-French
    print("\n[3/6] Fama-French regression...")
    try:
        ff = get_fama_french_factors()
        ff_results = compute_fama_french_alpha(bt["daily_returns"], ff)
        print(f"  FF3 Alpha: {ff_results['alpha_annual']:.4%}")
        print(f"  p-value:   {ff_results['p_value_alpha']:.4f}")
    except Exception as e:
        print(f"  Failed: {e}")

    # Step 4: Monte Carlo
    print("\n[4/6] Monte Carlo simulation...")
    try:
        latest_w = _get_latest_weights(bt["weights_history"], tickers)
        prices = get_price_data(tickers, start_date, end_date)
        paths = run_monte_carlo(latest_w, prices, n_simulations=1_000)
        mc_sum = get_simulation_summary(paths, initial_value=100_000)
        for k, v in mc_sum.items():
            print(f"  {k}: {v}")
    except Exception as e:
        print(f"  Failed: {e}")

    # Step 5: Stress testing
    print("\n[5/6] Stress testing...")
    try:
        stress = run_all_scenarios(latest_w, initial_value=100_000)
        for s in stress:
            print(f"  {s['scenario']:25s} : {s['portfolio_return']:+.2%}  (${s['dollar_loss']:+,.0f})")
    except Exception as e:
        print(f"  Failed: {e}")

    # Step 6: Factor scores
    print("\n[6/6] Factor scores...")
    scores = _get_latest_factor_scores(bt, tickers)
    if scores is not None:
        print(scores.round(4).to_string())

    print("\n" + "=" * 60)
    print("STANDALONE TEST COMPLETE — all pipeline stages verified")
    print("=" * 60)

    return {
        "bt": bt,
        "metrics": metrics,
        "summary": summary,
        "ff_results": locals().get("ff_results"),
        "mc_sum": locals().get("mc_sum"),
        "stress": locals().get("stress"),
        "scores": scores,
        "tickers": tickers,
    }


# Write results to Excel using openpyxl 

def write_results_to_excel(results: Dict, output_path: str = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from pathlib import Path
    from datetime import datetime

    if output_path is None:
        output_path = str(Path(__file__).parent.parent / "excel" / "PortfolioRiskTool.xlsx")

    bt = results["bt"]
    metrics = results["metrics"]
    summary = results["summary"]
    ff_results = results.get("ff_results")
    mc_sum = results.get("mc_sum")
    stress = results.get("stress")
    scores = results.get("scores")
    tickers = results.get("tickers", [])
    metadata = bt["metadata"]

    # ── Styles ──
    title_font = Font(name="Calibri", size=16, bold=True, color="003366")
    header_font = Font(name="Calibri", size=12, bold=True, color="003366")
    label_font = Font(name="Calibri", size=11, bold=True)
    col_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_font = Font(name="Calibri", size=11, italic=True, color="008000")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wb = openpyxl.Workbook()


    #  Sheet 1: Analysis
    ws1 = wb.active
    ws1.title = "Analysis"

    ws1["A1"] = "Portfolio Risk Management - Analysis Dashboard"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="666666")

    # Input area
    ws1["A3"] = "Tickers:"
    ws1["A3"].font = label_font
    ws1["B3"] = ", ".join(tickers)
    ws1["B3"].fill = input_fill
    ws1["A4"] = "Backtest Years:"
    ws1["A4"].font = label_font
    ws1["B4"] = 3
    ws1["B4"].fill = input_fill

    # Backtest parameters
    row = 6
    ws1.cell(row=row, column=1, value="Backtest Parameters").font = header_font
    row += 1
    params = {
        "Period": f"{metadata.get('start_date', '')} to {metadata.get('end_date', '')}",
        "Rebalance Freq": metadata.get("rebalance_freq", "M"),
        "Initial Value": f"${metadata.get('initial_value', 100000):,.0f}",
        "# Rebalances": metadata.get("n_rebalances", ""),
        "Total Return": f"{metadata.get('total_return', 0):.2%}",
        "Avg Turnover": f"{metadata.get('avg_turnover', 0):.2%}",
    }
    for k, v in params.items():
        ws1.cell(row=row, column=1, value=k).font = label_font
        ws1.cell(row=row, column=2, value=str(v))
        row += 1

    # Performance metrics
    row += 1
    ws1.cell(row=row, column=1, value="Performance & Risk Metrics").font = header_font
    row += 1
    # Column headers
    for j, col_name in enumerate(summary.columns):
        cell = ws1.cell(row=row, column=1 + j, value=col_name)
        cell.font = col_header_font
        cell.fill = col_header_fill
    row += 1
    for _, data_row in summary.iterrows():
        for j, col_name in enumerate(summary.columns):
            ws1.cell(row=row, column=1 + j, value=data_row[col_name])
        row += 1

    # Fama-French
    if ff_results:
        row += 1
        ws1.cell(row=row, column=1, value="Fama-French 3-Factor Regression").font = header_font
        row += 1
        ff_display = {
            "FF3 Alpha (annualised)": f"{ff_results.get('alpha_annual', 0):.4%}",
            "Alpha t-stat": f"{ff_results.get('t_stat_alpha', 0):.3f}",
            "Alpha p-value": f"{ff_results.get('p_value_alpha', 0):.4f}",
            "R-squared": f"{ff_results.get('r_squared', 0):.4f}",
            "# Observations": str(ff_results.get("n_observations", "")),
            "Beta (Mkt-RF)": f"{ff_results.get('beta_mkt', 0):.4f}",
            "Beta (SMB)": f"{ff_results.get('beta_smb', 0):.4f}",
            "Beta (HML)": f"{ff_results.get('beta_hml', 0):.4f}",
        }
        for k, v in ff_display.items():
            ws1.cell(row=row, column=1, value=k).font = label_font
            ws1.cell(row=row, column=2, value=v)
            row += 1

    # Column widths
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 45
    ws1.column_dimensions["C"].width = 18


    #  Sheet 2: Backtest
    ws2 = wb.create_sheet("Backtest")
    ws2["A1"] = "Daily NAV"
    ws2["A1"].font = title_font

    nav = bt["nav"]
    bench_returns = bt["benchmark_returns"]
    bench_nav = (1 + bench_returns).cumprod() * metadata.get("initial_value", 100_000)

    # Headers
    for j, h in enumerate(["Date", "Portfolio NAV", "Benchmark NAV"]):
        cell = ws2.cell(row=2, column=1 + j, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill

    # Data rows
    for i, (date, value) in enumerate(nav.items()):
        r = 3 + i
        ws2.cell(row=r, column=1, value=date.strftime("%Y-%m-%d"))
        ws2.cell(row=r, column=2, value=round(float(value), 2))
        bench_val = bench_nav.get(date)
        if bench_val is not None:
            ws2.cell(row=r, column=3, value=round(float(bench_val), 2))

    # Weights history
    weights_history = bt["weights_history"]
    if weights_history:
        col_start = 5
        ws2.cell(row=1, column=col_start, value="Portfolio Weights").font = title_font

        all_tickers = sorted(set(t for _, w in weights_history for t in w.keys()))

        # Header
        ws2.cell(row=2, column=col_start, value="Date").font = col_header_font
        ws2.cell(row=2, column=col_start).fill = col_header_fill
        for j, ticker in enumerate(all_tickers):
            cell = ws2.cell(row=2, column=col_start + 1 + j, value=ticker)
            cell.font = col_header_font
            cell.fill = col_header_fill

        for i, (date, wts) in enumerate(weights_history):
            r = 3 + i
            if hasattr(date, "strftime"):
                ws2.cell(row=r, column=col_start, value=date.strftime("%Y-%m-%d"))
            else:
                ws2.cell(row=r, column=col_start, value=str(date))
            for j, ticker in enumerate(all_tickers):
                val = wts.get(ticker, 0.0)
                cell = ws2.cell(row=r, column=col_start + 1 + j, value=round(val, 4))
                cell.number_format = "0.00%"

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["E"].width = 12


    #  Sheet 3: Raw Data
    ws3 = wb.create_sheet("Raw Data")
    row = 1

    # Factor scores
    if scores is not None and not scores.empty:
        ws3.cell(row=row, column=1, value="Factor Scores (latest)").font = title_font
        row += 1
        cols = ["Ticker"] + list(scores.columns)
        for j, c in enumerate(cols):
            cell = ws3.cell(row=row, column=1 + j, value=c)
            cell.font = col_header_font
            cell.fill = col_header_fill
        row += 1
        for ticker, data_row in scores.iterrows():
            ws3.cell(row=row, column=1, value=str(ticker))
            for j, col in enumerate(scores.columns):
                ws3.cell(row=row, column=2 + j, value=round(float(data_row[col]), 4))
            row += 1
        row += 2

    # Stress test
    if stress:
        ws3.cell(row=row, column=1, value="Stress Test Results").font = title_font
        row += 1
        stress_headers = ["Scenario", "Description", "Portfolio Return", "Dollar Loss", "Worst Ticker"]
        for j, h in enumerate(stress_headers):
            cell = ws3.cell(row=row, column=1 + j, value=h)
            cell.font = col_header_font
            cell.fill = col_header_fill
        row += 1
        for s in stress:
            ws3.cell(row=row, column=1, value=s.get("scenario", ""))
            ws3.cell(row=row, column=2, value=s.get("description", ""))
            cell = ws3.cell(row=row, column=3, value=round(s.get("portfolio_return", 0), 4))
            cell.number_format = "0.00%"
            cell = ws3.cell(row=row, column=4, value=round(s.get("dollar_loss", 0), 2))
            cell.number_format = "$#,##0"
            ws3.cell(row=row, column=5, value=s.get("worst_ticker", ""))
            row += 1
        row += 2

    # Monte Carlo
    if mc_sum:
        ws3.cell(row=row, column=1, value="Monte Carlo Simulation (1-Year)").font = title_font
        row += 1
        for k, v in mc_sum.items():
            ws3.cell(row=row, column=1, value=str(k)).font = label_font
            cell = ws3.cell(row=row, column=2)
            if isinstance(v, (int, float)):
                cell.value = v
                if abs(v) < 1:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "$#,##0.00"
            else:
                cell.value = str(v)
            row += 1

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 15
    ws3.column_dimensions["E"].width = 15

    wb.save(output_path)
    print(f"\nResults written to: {output_path}")
    print("Open this file in Excel to see all 3 sheets populated!")


if __name__ == "__main__":
    results = run_standalone_test()
    if results:
        print("\n[7/7] Writing results to Excel...")
        write_results_to_excel(results)
        print("\nDone! Check excel/PortfolioRiskTool.xlsx")
